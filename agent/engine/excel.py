import openpyxl
from common.database import SessionLocal
from agent.models.syuppin_item import *
from django.http import HttpResponse


def download_excel():
    
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=%s' % 'test_xlm.xls'
    wb = openpyxl.load_workbook("test_xlm.xlsx")
    ws = wb["テンプレート"]

    db = SessionLocal()

    syuppin_item = db.query(SyuppinItem).all()
    row = 3
    for data in syuppin_item:
        row += 1
        ws.cell(row=row, column=4, value=data.item_name)
        ws.cell(row=row, column=11, value=data.price)
        ws.cell(row=row, column=12, value=data.image_url1)

    wb.save(response)
    return response