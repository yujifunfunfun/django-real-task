from ..models.setting_table import *
from django_tables2 import SingleTableView
from ..tables.setting_table import *
from ..models.setting_table import *
from ..models.syuppin_item import *
import openpyxl
from django.http import HttpResponse


class SettingTableView(SingleTableView):
    table_class = ItemTable
    template_name = 'setting/setting_table.html'
    def get_queryset(self):
        return SyuppinItem.objects.all()

    def post(self,request):

        response = HttpResponse(content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'attachment; filename=%s' % 'test_xlm.xlsm'
        wb = openpyxl.load_workbook("setting/views/test_xlm.xlsm",keep_vba=True)
        ws = wb["テンプレート"]

        syuppin_item = SyuppinItem.objects.all()

        row = 3

        for data in syuppin_item:
            row += 1
            ws.cell(row=row, column=4, value=data.item_name)
            ws.cell(row=row, column=11, value=data.price)
            ws.cell(row=row, column=12, value=data.image_url1)

        wb.save(response)
        wb.close()
        return response


    